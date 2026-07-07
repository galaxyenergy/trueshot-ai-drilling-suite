from pathlib import Path
from copy import copy
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook


TEMPLATE_DIR = Path("templates") / "surveys"

TRUESHOT_TEMPLATE = TEMPLATE_DIR / "trueshot_survey_template.xlsm"
OXY_TEMPLATE = TEMPLATE_DIR / "oxy_survey_template.xlsx"


def _safe_number(value, default=0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def _copy_row_style(ws, source_row, target_row, max_col):
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)

        if src.has_style:
            dst._style = copy(src._style)

        if src.number_format:
            dst.number_format = src.number_format

        if src.alignment:
            dst.alignment = copy(src.alignment)

        if src.font:
            dst.font = copy(src.font)

        if src.fill:
            dst.fill = copy(src.fill)

        if src.border:
            dst.border = copy(src.border)


def read_trueshot_survey_template(template_path=TRUESHOT_TEMPLATE):
    """
    Reads survey values from TRUEshot / Primo survey format.

    TRUEshot sheet:
    B = MD
    C = Inclination
    D = Azimuth
    E = Course Length
    F = TVD
    G = Vertical Section
    H/I = N/S value and direction
    J/K = E/W value and direction
    L = Closure Distance
    M = Closure Direction
    N = DLS
    O = Build Rate
    P = Walk Rate
    """

    if not Path(template_path).exists():
        raise FileNotFoundError(f"TRUEshot template not found: {template_path}")

    wb = load_workbook(
        template_path,
        data_only=True,
        keep_vba=True
    )

    ws = wb["Surveys"] if "Surveys" in wb.sheetnames else wb.active

    rows = []

    blank_count = 0

    for row in range(10, ws.max_row + 1):
        md = ws.cell(row, 2).value

        if md is None or md == "":
            blank_count += 1
            if blank_count >= 10:
                break
            continue

        blank_count = 0

        survey_label = ws.cell(row, 1).value

        survey_type = "MWD"
        if isinstance(survey_label, str) and "tie" in survey_label.lower():
            survey_type = "Tie In"

        rows.append(
            {
                "Type": survey_type,
                "Survey_Number": survey_label,
                "MD": _safe_number(ws.cell(row, 2).value),
                "Inc": _safe_number(ws.cell(row, 3).value),
                "Azi": _safe_number(ws.cell(row, 4).value),
                "Course_Length": _safe_number(ws.cell(row, 5).value),
                "TVD": _safe_number(ws.cell(row, 6).value),
                "Vertical_Section": _safe_number(ws.cell(row, 7).value),
                "NS": _safe_number(ws.cell(row, 8).value),
                "NS_Dir": ws.cell(row, 9).value or "N",
                "EW": _safe_number(ws.cell(row, 10).value),
                "EW_Dir": ws.cell(row, 11).value or "W",
                "Closure_Distance": _safe_number(ws.cell(row, 12).value),
                "Closure_Direction": _safe_number(ws.cell(row, 13).value),
                "DLS": _safe_number(ws.cell(row, 14).value),
                "Build_Rate": _safe_number(ws.cell(row, 15).value),
                "Walk_Rate": _safe_number(ws.cell(row, 16).value),
            }
        )

    return pd.DataFrame(rows)


def create_oxy_survey_file(survey_df, output_name="generated_oxy_survey.xlsx"):
    """
    Creates Oxy client survey format from survey dataframe.
    """

    if not OXY_TEMPLATE.exists():
        raise FileNotFoundError(f"Oxy template not found: {OXY_TEMPLATE}")

    wb = load_workbook(OXY_TEMPLATE)
    ws = wb.active

    start_row = 20
    max_col = 14

    # Clear old survey table area
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None

    # Populate survey rows
    for idx, record in survey_df.reset_index(drop=True).iterrows():
        excel_row = start_row + idx

        if excel_row > 21:
            _copy_row_style(ws, 21, excel_row, max_col)

        survey_type = record.get("Type", "MWD")

        if idx == 0:
            survey_type = "Tie In"
        elif not survey_type or str(survey_type).lower() == "nan":
            survey_type = "MWD"

        ws.cell(excel_row, 1).value = survey_type
        ws.cell(excel_row, 2).value = record.get("MD", 0)
        ws.cell(excel_row, 3).value = record.get("Inc", 0)
        ws.cell(excel_row, 4).value = record.get("Azi", 0)
        ws.cell(excel_row, 5).value = record.get("Course_Length", 0)
        ws.cell(excel_row, 6).value = record.get("TVD", 0)
        ws.cell(excel_row, 7).value = record.get("Vertical_Section", 0)
        ws.cell(excel_row, 8).value = record.get("NS", 0)
        ws.cell(excel_row, 9).value = record.get("NS_Dir", "N")
        ws.cell(excel_row, 10).value = record.get("EW", 0)
        ws.cell(excel_row, 11).value = record.get("EW_Dir", "W")
        ws.cell(excel_row, 12).value = record.get("Closure_Distance", 0)
        ws.cell(excel_row, 13).value = record.get("Closure_Direction", 0)
        ws.cell(excel_row, 14).value = record.get("DLS", 0)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def create_trueshot_survey_file(survey_df, output_name="generated_trueshot_survey.xlsm"):
    """
    Creates TRUEshot survey format.

    Important:
    This preserves the .xlsm template and only writes MD, Inclination, and Azimuth.
    Excel formulas in the TRUEshot template calculate TVD, VS, closure, DLS, build rate, and walk rate.
    """

    if not TRUESHOT_TEMPLATE.exists():
        raise FileNotFoundError(f"TRUEshot template not found: {TRUESHOT_TEMPLATE}")

    wb = load_workbook(
        TRUESHOT_TEMPLATE,
        keep_vba=True
    )

    ws = wb["Surveys"] if "Surveys" in wb.sheetnames else wb.active

    start_row = 10

    # Clear old MD / Inc / Azi input area only
    for row in range(start_row, ws.max_row + 1):
        ws.cell(row, 1).value = None
        ws.cell(row, 2).value = None
        ws.cell(row, 3).value = None
        ws.cell(row, 4).value = None

    # Populate TRUEshot input columns
    for idx, record in survey_df.reset_index(drop=True).iterrows():
        excel_row = start_row + idx

        if excel_row > 11:
            _copy_row_style(ws, 11, excel_row, 16)

        if idx == 0:
            ws.cell(excel_row, 1).value = "Tie In"
        else:
            ws.cell(excel_row, 1).value = idx

        ws.cell(excel_row, 2).value = record.get("MD", 0)
        ws.cell(excel_row, 3).value = record.get("Inc", 0)
        ws.cell(excel_row, 4).value = record.get("Azi", 0)

    # Ask Excel to recalculate formulas when file opens
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def generate_survey_files_from_templates():
    """
    Reads the TRUEshot / Primo template survey data,
    then creates both downloadable survey files.
    """

    survey_df = read_trueshot_survey_template()

    trueshot_file = create_trueshot_survey_file(survey_df)
    oxy_file = create_oxy_survey_file(survey_df)

    return {
        "survey_df": survey_df,
        "trueshot_file": trueshot_file,
        "oxy_file": oxy_file,
    }
    
    